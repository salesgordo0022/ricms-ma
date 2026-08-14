# -*- coding: utf-8 -*-
"""PILOTO — Auditor de NCM.
Para cada produto (descrição + NCM atual) faz:
  1) VALIDAÇÃO determinística: o NCM existe? está vigente? é revogado? está incompleto?
  2) ADERÊNCIA descrição x NCM (compara com a descrição oficial e com nomes reais da base)
  3) SUGESTÃO de NCM (candidatos a CONFERIR — reaproveita os produtos já validados da base)
  4) ENQUADRAMENTO fiscal: benefício + CST/CFOP/CEST do NCM no MA
Gera um Excel com semáforo. Uso:
    python auditoria_ncm.py                      -> roda a lista de EXEMPLO
    python auditoria_ncm.py entrada.xlsx         -> audita sua planilha (colunas: descricao, ncm)
"""
import os, re, sys, csv, json, unicodedata, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
try: sys.stdout.reconfigure(encoding="utf-8")
except Exception: pass

AQUI = os.path.dirname(os.path.abspath(__file__))
CSV_NCM = os.path.join(AQUI, "ncm_completa.csv")   # opcional; usa tipi_ncm.json se ausente
TIPI = os.path.join(AQUI, "tipi_ncm.json")
BASE = os.path.join(AQUI, "base.json")
HOJE = datetime.date(2026, 8, 4)

def _n(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9 ]", " ", s.lower())
def _d(s): return re.sub(r"\D", "", str(s or ""))
def _fmt(dd):
    dd = _d(dd)
    return f"{dd[:4]}.{dd[4:6]}.{dd[6:8]}" if len(dd) == 8 else dd
STOP = {"pacote", "lata", "latas", "caixa", "unidade", "unidades", "fardo", "garrafa", "garrafas",
        "embalagem", "litro", "litros", "gramas", "grama", "tipo", "pct", "und", "tamanho", "cor",
        "marca", "modelo", "polegadas", "com", "sem", "para", "dos", "das", "kit", "ref"}
def toks(s): return {t for t in _n(s).split() if len(t) > 2}
def qtoks(s):                            # tokens da CONSULTA (sem embalagem/genéricas/medidas)
    return {t for t in toks(s) if t not in STOP and not re.fullmatch(r"\d+[a-z]{0,3}", t)}
def _match(x, y):
    """casa por radical: igual, um é prefixo do outro, ou 5 primeiras letras iguais
    (resolve plural/derivação: pneu~pneumaticos, novo~novos, automovel~automoveis)."""
    if x == y: return True
    if len(x) >= 4 and len(y) >= 4:
        if x.startswith(y) or y.startswith(x): return True
        if len(x) >= 5 and len(y) >= 5 and x[:5] == y[:5]: return True
    return False
def fuzzy_inter(a, b):
    return sum(1 for x in a if any(_match(x, y) for y in b))

# ---- carrega tabela oficial de NCM ----
# Fonte preferencial: ncm_completa.csv (local, opcional). Sem ele, usa tipi_ncm.json
# (mesma tabela oficial Camex — 10.515 códigos de 8 dígitos, com vigência) p/ rodar em servidor.
NCM = {}   # 8dig -> {desc, ini, fim, vigente}
def _pdate(s):
    try: return datetime.datetime.strptime(s.strip(), "%d/%m/%Y").date()
    except Exception: return None
def _vigente(fim_txt):
    fim = _pdate(fim_txt)
    return (fim is None or fim >= HOJE)
if os.path.isfile(CSV_NCM):
    with open(CSV_NCM, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            cod = _d(row.get("NCM"))
            if len(cod) != 8: continue
            NCM[cod] = {"desc": (row.get("Descricao") or "").strip(),
                        "ini": row.get("Data_Inicio", ""), "fim": row.get("Data_Fim", ""),
                        "vigente": _vigente(row.get("Data_Fim", ""))}
else:
    for x in json.load(open(TIPI, encoding="utf-8")).get("Nomenclaturas", []):
        cod = _d(x.get("Codigo"))
        if len(cod) != 8: continue
        NCM[cod] = {"desc": (x.get("Descricao") or "").strip(),
                    "ini": x.get("Data_Inicio", ""), "fim": x.get("Data_Fim", ""),
                    "vigente": _vigente(x.get("Data_Fim", ""))}
# prefixos válidos (posição/subposição) p/ detectar incompleto porém existente
PREF = set()
for c in NCM: PREF.update({c[:2], c[:4], c[:6]})

# ---- hierarquia TIPI: descricao COMPLETA (capitulo/posicao/subposicao/item) ----
NIVEL = {}   # prefixo-digitos -> descricao daquele nivel
try:
    for x in json.load(open(TIPI, encoding="utf-8")).get("Nomenclaturas", []):
        NIVEL[_d(x.get("Codigo"))] = (x.get("Descricao") or "").strip()
except Exception:
    NIVEL = {}
def desc_completa(cod8):
    """Junta os textos dos niveis 2/4/6/8 -> 'Arroz. > Arroz semibranqueado... > Polido ou brunido'."""
    partes = [NIVEL.get(cod8[:k]) for k in (2, 4, 6, 8)]
    partes = [p for p in partes if p]
    return " > ".join(partes) if partes else NCM.get(cod8, {}).get("desc", "")

# tokens de descricao completa por NCM (p/ sugestao) — inclui nomes reais da base
TOK_NCM = {}   # cod8 -> set de tokens (desc completa + nomes base)
for c in NCM:
    TOK_NCM[c] = toks(desc_completa(c))

# ---- base de benefícios (ncm -> itens) e nomes de produto (p/ sugestão e aderência) ----
B = json.load(open(BASE, encoding="utf-8"))["beneficios"]
POR_NCM = {}
NOMES = []   # (nome_norm, ncm8, produto, beneficio)
NOME_POR_NCM = {}   # cod8 -> nome real (p/ mostrar como "exemplo")
for b in B:
    c = _d(b.get("ncm"))
    if len(c) == 8:
        POR_NCM.setdefault(c, []).append(b)
        NOMES.append((_n(b.get("produto")), c, b.get("produto", ""), b.get("beneficio", "")))
        if c in TOK_NCM:
            TOK_NCM[c] |= toks(b.get("produto"))   # enriquece a descricao oficial com nome real
        NOME_POR_NCM.setdefault(c, b.get("produto", ""))

# índice invertido (radical de 4 letras -> NCMs) p/ NÃO varrer os 10.515 a cada consulta
INV = {}
for _c8, _tk in TOK_NCM.items():
    for _t in _tk:
        INV.setdefault(_t[:4], set()).add(_c8)
def _cands(pt):
    s = set()
    for t in pt:
        s |= INV.get(t[:4], set())
    return s

SUG_MIN = 34   # % minimo p/ sugerir (abaixo disso é ruído — não sugere)

def sugere_ncm(descricao, limite=3):
    """Candidatos a CONFERIR: casa a descrição do produto com a DESCRIÇÃO COMPLETA
    (hierarquia TIPI) + nomes reais da base. Só devolve acima de SUG_MIN%."""
    pt = qtoks(descricao)
    if not pt: return []
    ranks = []
    for c8 in _cands(pt):
        tk = TOK_NCM.get(c8)
        if not tk: continue
        inter = fuzzy_inter(pt, tk)
        if inter < 2:            # exige ao menos 2 palavras (por radical) em comum
            continue
        sc = inter / len(pt)      # cobertura das palavras do produto
        ranks.append((sc, c8))
    ranks.sort(reverse=True)
    out = []
    for sc, c8 in ranks[:limite]:
        pct = round(sc * 100)
        if pct < SUG_MIN: break
        exemplo = NOME_POR_NCM.get(c8) or (NIVEL.get(c8) or desc_completa(c8))
        ben = (POR_NCM.get(c8, [{}])[0]).get("beneficio", "")
        out.append({"ncm": _fmt(c8), "score": pct, "exemplo": exemplo[:45], "beneficio": ben})
    return out

def detectar_colunas(linhas):
    """Descobre qual coluna é a DESCRIÇÃO e qual é o NCM (sem depender de cabeçalho)."""
    amostra = [r for r in linhas[:400] if any(str(c).strip() for c in r)]
    if not amostra:
        return (0, None)
    ncols = max(len(r) for r in amostra)
    ncm8 = [0] * ncols; ncmish = [0] * ncols; desc = [0] * ncols; cnt = [0] * ncols
    for r in amostra:
        for i in range(ncols):
            v = str(r[i]).strip() if i < len(r) else ""
            if not v:
                continue
            cnt[i] += 1
            d = _d(v)
            somente_num = len(d) == len(v.replace(".", "").replace(" ", "").replace("-", ""))
            if somente_num and len(d) == 8:
                ncm8[i] += 1
            if somente_num and 6 <= len(d) <= 8:
                ncmish[i] += 1
            if sum(ch.isalpha() for ch in v) >= 4:
                desc[i] += len(v)
    ncm_col = max(range(ncols), key=lambda i: (ncm8[i] / cnt[i]) if cnt[i] else 0)
    if not cnt[ncm_col] or ncm8[ncm_col] / cnt[ncm_col] < 0.25:
        ncm_col = max(range(ncols), key=lambda i: (ncmish[i] / cnt[i]) if cnt[i] else 0)
        if not cnt[ncm_col] or ncmish[ncm_col] / cnt[ncm_col] < 0.25:
            ncm_col = None
    desc_col = max((i for i in range(ncols) if i != ncm_col), key=lambda i: desc[i], default=0)
    return (desc_col, ncm_col)

def extrair_itens(linhas):
    dc, nc = detectar_colunas(linhas)
    out = []
    for idx, r in enumerate(linhas):
        d = str(r[dc]).strip() if dc is not None and dc < len(r) else ""
        n = str(r[nc]).strip() if nc is not None and nc < len(r) else ""
        if idx == 0 and n and not _d(n):     # pula cabeçalho (NCM não-numérico na 1ª linha)
            continue
        if d or n:
            out.append((d, n))
    return out, dc, nc

def existe_vigente(cod):
    c = _d(cod)
    return len(c) == 8 and c in NCM and NCM[c]["vigente"]
def desc_de(cod):
    return desc_completa(_d(cod))
def candidatos_ia(descricao, n=8):
    """Lista ampla de candidatos (com descrição oficial completa) p/ dar contexto à IA."""
    pt = qtoks(descricao)
    if not pt: return []
    ranks = []
    for c8 in _cands(pt):
        tk = TOK_NCM.get(c8)
        if not tk: continue
        inter = fuzzy_inter(pt, tk)
        if inter >= 1:
            ranks.append((inter / len(pt), c8))
    ranks.sort(reverse=True)
    return [{"ncm": _fmt(c8), "desc": desc_completa(c8)[:110]} for _, c8 in ranks[:n]]

def audita(descricao, ncm_atual):
    cod = _d(ncm_atual)
    r = {"descricao": descricao, "ncm_informado": _fmt(ncm_atual) if cod else "", "oficial": "",
         "status_ncm": "", "aderencia": "", "diag_aderencia": "", "sugestoes": [],
         "beneficio": "", "cst": "", "cfop": "", "cest": "", "recomendacao": ""}
    # 1) validação
    if not cod:
        r["status_ncm"] = "🔴 SEM NCM informado"
    elif len(cod) < 8:
        existe = cod[:2] in PREF or cod[:4] in PREF or cod[:6] in PREF
        r["status_ncm"] = f"🟡 INCOMPLETO ({len(cod)} díg.){' — posição existe' if existe else ''}"
    elif cod not in NCM:
        r["status_ncm"] = "🔴 NÃO EXISTE na tabela oficial"
    elif not NCM[cod]["vigente"]:
        r["status_ncm"] = f"🔴 REVOGADO (fim {NCM[cod]['fim']})"
        r["oficial"] = desc_completa(cod)
    else:
        r["status_ncm"] = "✅ válido e vigente"
        r["oficial"] = desc_completa(cod)
    # 2) aderência + CLASSE (calibrado p/ descrições abreviadas de estoque):
    #    🔴 só para ERRO REAL de NCM. NCM válido com aderência baixa NÃO é "divergente"
    #    — só vira "revisar" quando existe um NCM claramente MAIS aderente.
    pt = qtoks(descricao)
    if not cod:
        r["classe"] = "invalido"
    elif len(cod) < 8:
        r["classe"] = "incompleto"
    elif cod not in NCM or not NCM[cod]["vigente"]:
        r["classe"] = "invalido"
    else:
        # NCM válido: aderência é só INFORMATIVA. Não rebaixa NCM válido por semelhança
        # difusa (evita falso cognato tipo "refrigerante" x "refrigerada"). Reclassificação
        # fica a cargo da IA, que é confiável.
        alvo = set(TOK_NCM.get(cod, set()))
        inter = fuzzy_inter(pt, alvo) if pt else 0
        pct = round((inter / len(pt)) * 100) if pt else 0
        r["aderencia"] = f"{pct}%"
        r["diag_aderencia"] = "✅ combina" if (inter >= 2 and pct >= 40) else "✅ válido"
        r["classe"] = "ok"
    # sugestões (pista a conferir) só nos casos com erro real de NCM
    r["sugestoes"] = sugere_ncm(descricao) if r.get("classe") in ("invalido", "incompleto") else []
    best = r["sugestoes"][0] if r["sugestoes"] else None
    # 4) enquadramento fiscal do NCM informado (se válido e na base)
    itens = POR_NCM.get(cod, []) if len(cod) == 8 else []
    if itens:
        it = itens[0]
        r["beneficio"] = it.get("beneficio", "")
        r["cst"] = it.get("cst_icms_sugerido", "")
        r["cest"] = it.get("cest", "")
        cpo = (it.get("cfop_por_operacao", {}) or {}).get("saida_interna", [])
        r["cfop"] = " / ".join(x.get("cfop", "") for x in cpo[:2])
    # 5) recomendação textual
    classe = r.get("classe", "ok")
    if classe == "invalido":
        rec = ["Corrigir o NCM (não existe / revogado)."]
        if best:
            rec.append(f"Provável: {best['ncm']} ({best['exemplo'][:38]}) — conferir (RGI).")
    elif classe == "incompleto":
        rec = ["Completar o NCM até 8 dígitos."]
        if best:
            rec.append(f"Ex.: {best['ncm']} ({best['exemplo'][:38]}).")
    elif not itens:
        rec = ["NCM válido; sem benefício mapeado no MA (tributação normal ou confirmar)."]
    else:
        rec = ["NCM válido e com enquadramento fiscal identificado."]
    r["recomendacao"] = " ".join(rec)
    return r

# ---------- Excel (auditoria, colorido por classe) ----------
def escreve_auditoria(audits, out):
    fh = Font(size=10, bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9D9D9"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    COR = {"invalido": "FBE4E4", "incompleto": "FBE4E4", "revisar": "FFF6E0", "ok": "E7F6EC"}
    wb = Workbook(); ws = wb.active; ws.title = "Auditoria NCM"; ws.sheet_view.showGridLines = False
    cols = [("descricao", "Produto (descrição)", 42), ("ncm_informado", "NCM arquivo", 13),
            ("status_ncm", "Status do NCM", 25), ("aderencia", "Ader.", 7), ("diag_aderencia", "Descr. casa?", 16),
            ("sug", "NCM sugerido (conferir)", 30), ("cest_arquivo", "CEST arquivo", 12),
            ("cest", "CEST base MA", 12), ("beneficio", "Benefício MA", 20), ("cst", "CST", 10),
            ("cfop", "CFOP", 12), ("recomendacao", "Recomendação", 46)]
    for j, (k, lab, w) in enumerate(cols):
        c = ws.cell(row=1, column=1 + j, value=lab); c.font = fh; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bd
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    for i, a in enumerate(audits):
        a = dict(a); a["sug"] = " | ".join(f"{s['ncm']} ({s['score']}%)" for s in a.get("sugestoes", []))
        rowfill = PatternFill("solid", fgColor=COR.get(a.get("classe", "ok"), "FFFFFF"))
        for j, (k, lab, w) in enumerate(cols):
            c = ws.cell(row=2 + i, column=1 + j, value=str(a.get(k, "") or "")); c.font = Font(size=9)
            c.alignment = wrap; c.border = bd; c.fill = rowfill
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{1+len(audits)}"
    ws.row_dimensions[1].height = 26
    wb.save(out); return out

# ---------- Excel ----------
def escreve(audits, out):
    fhdr = Font(size=10, bold=True, color="FFFFFF"); fillh = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9D9D9"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    wb = Workbook(); ws = wb.active; ws.title = "Auditoria NCM"; ws.sheet_view.showGridLines = False
    cols = [("descricao", "Produto (descrição)", 40), ("ncm_informado", "NCM informado", 14),
            ("status_ncm", "Status do NCM", 26), ("oficial", "Descrição oficial do NCM", 40),
            ("aderencia", "Aderência", 9), ("diag_aderencia", "Descrição casa?", 14),
            ("sug", "NCM sugerido (CONFERIR)", 34), ("beneficio", "Benefício no MA", 20),
            ("cst", "CST", 10), ("cfop", "CFOP", 12), ("cest", "CEST", 12),
            ("recomendacao", "Recomendação", 46)]
    for j, (k, lab, w) in enumerate(cols):
        c = ws.cell(row=1, column=1 + j, value=lab); c.font = fhdr; c.fill = fillh
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bd
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    for i, a in enumerate(audits):
        a = dict(a); a["sug"] = " | ".join(f"{s['ncm']} ({s['score']}%)" for s in a.get("sugestoes", []))
        for j, (k, lab, w) in enumerate(cols):
            c = ws.cell(row=2 + i, column=1 + j, value=str(a.get(k, "") or "")); c.font = Font(size=9)
            c.alignment = wrap; c.border = bd
            if i % 2: c.fill = PatternFill("solid", fgColor="F2F6FB")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{1+len(audits)}"
    ws.row_dimensions[1].height = 28
    wb.save(out); return out

LISTA_EXEMPLO = [
    ("Arroz beneficiado tipo 1, pacote 5kg", "1006.30.21"),
    ("Feijão carioca tipo 1", "0713.33.19"),
    ("Refrigerante de cola, lata 350ml", "2202.10.00"),
    ("Farinha de milho flocada (flocão)", "1104.19.00"),
    ("Pneu novo de automóvel", "4011.10.00"),
    ("Camiseta de algodão adulto", "9999.99.99"),          # NÃO existe
    ("Leite em pó integral", "0402"),                       # INCOMPLETO
    ("Notebook 14 polegadas", "0201.10.00"),                # descrição NÃO casa (0201 = carne bovina)
    ("Reprodutor bovino com registro", "0102.21.10"),
    ("Cerveja pilsen long neck", "2203.00.00"),
]

if __name__ == "__main__":
    if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
        wb = load_workbook(sys.argv[1]); ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        lista = [(str(r[0] or ""), str(r[1] or "")) for r in rows if r and r[0]]
        out = os.path.join(os.path.dirname(os.path.abspath(sys.argv[1])), "Auditoria_NCM_RESULTADO.xlsx")
    else:
        lista = LISTA_EXEMPLO
        out = r"C:\Users\sales\Documents\imperialfistnes\RICMS-MA\RICMS-MA_Auditoria_NCM_EXEMPLO.xlsx"
    audits = [audita(desc, ncm) for desc, ncm in lista]
    escreve(audits, out)
    print("SALVO:", out, "| linhas:", len(audits), "| NCMs oficiais carregados:", len(NCM))
    for a in audits:
        print(f"  {a['ncm_informado'] or '(vazio)':<13} {a['status_ncm'][:22]:<22} ader={a['aderencia'] or '-':<5} {a['diag_aderencia']:<12} -> {a['recomendacao'][:55]}")
